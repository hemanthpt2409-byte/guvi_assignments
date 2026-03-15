from openpyxl import load_workbook
from datetime import datetime

class ExcelUtils:

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(file)
        self.sheet = self.wb.active

    def get_data(self):
        data = []

        for row in range(2, self.sheet.max_row + 1):
            username = self.sheet.cell(row, 2).value
            password = self.sheet.cell(row, 3).value

            data.append((username, password, row))

        return data

    def write_result(self, row, result):
        now = datetime.now()

        self.sheet.cell(row, 4).value = now.date()
        self.sheet.cell(row, 5).value = now.time().strftime("%H:%M:%S")
        self.sheet.cell(row, 7).value = result

        self.wb.save(self.file)